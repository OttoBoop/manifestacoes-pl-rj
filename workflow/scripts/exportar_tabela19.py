#!/usr/bin/env python3
"""Acrescenta os despachos CONCLUÍDOS do time à tabela 19 ("Manifestação PLs SUBDEI")
do Projetos SUBDEI_2025-2028.xlsx, no formato da própria tabela:

    Ação (ofício) | Prazo | Responsável | Status | Observação (PL nº X, que "EMENTA")

Responsável: "Otávio" + quem fez ALTERAÇÃO na peça como coautor (padrão da casa
"Marcel + Luiza"). Aprovação sem alteração não vira coautoria.
Fonte: controle-excel/revisao_map.csv (status Concluído) + ementas oficiais
extraídas das próprias peças. PL 1884 vira duas linhas (manifestação + autógrafo),
seguindo o padrão de uma linha por ofício da tabela.
"""

import csv
import glob
import os
import re
import shutil
import sys
from copy import copy

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
MAP = os.path.join(BASE, "controle-excel", "revisao_map.csv")
TARGET = "/home/otavio/Downloads/Projetos SUBDEI_2025-2028.xlsx"
BACKUP = "/home/otavio/Downloads/Projetos SUBDEI_2025-2028_backup_2026-07-16.xlsx"

# Ofício por PL (fonte: INDICE-AVALIACOES.md), já no padrão da tabela 19
OFICIOS = {
    ("PL", "74"): "CVL-OFI-2026/00823", ("PL", "1795"): "SMG-OFI-2026/01526",
    ("PL", "1826"): "SMG-OFI-2026/01665", ("PL", "1840"): "CVL-OFI-2026/00925",
    ("PL", "1844"): "CVL-OFI-2026/00992", ("PL", "1866"): "CVL-OFI-2026/01133",
    ("PL", "1883"): "CVL-OFI-2026/01214", ("PL", "1884"): "CVL-OFI-2026/01218",
    ("PL", "1897"): "CVL-OFI-2026/01276", ("PL", "1900"): "CVL-OFI-2026/01294",
    ("PL", "1904"): "CVL-OFI-2026/01342", ("PL", "1921"): "CVL-OFI-2026/01476",
    ("PL", "1934"): "CVL-OFI-2026/01568", ("PL", "1954"): "CVL-OFI-2026/01722",
    ("PLC", "98"): "CVL-OFI-2026/01227", ("PLC", "102"): "CVL-OFI-2026/01268",
    ("PLC", "104"): "CVL-OFI-2026/01283", ("PLC", "105"): "CVL-OFI-2026/01322",
    ("PLC", "106"): "CVL-OFI-2026/01329", ("PL", "163"): "Despacho de Ciência (SEI 3412879)",
    ("PL", "1320"): "Despacho de Ciência (SEI 3412944)", ("PL", "1538"): "SGV-OFI-2025/10541",
    ("PL", "1964"): "CVL-OFI-2026/01765", ("PL", "1985"): "CVL-OFI-2026/01902",
    ("PL", "1990"): "CVL-OFI-2026/01933", ("PL", "2007"): "CVL-OFI-2026/02002",
    ("PLC", "112"): "CVL-OFI-2026/02108", ("PL", "2040"): "CVL-OFI-2026/02229",
    ("PELOM", "5"): "CVL-OFI-2026/02288", ("PL", "2076"): "CVL-OFI-2026/03513",
    ("PL", "2078"): "CVL-OFI-2026/03548", ("PL", "2097"): "CVL-OFI-2026/03646",
    ("PL", "799"): "CVL-OFI-2026/00934", ("PL", "2261"): "CVL-OFI-2026/04434",
    ("PL", "2262"): "CVL-OFI-2026/04436", ("PL", "2263"): "CVL-OFI-2026/04439",
    ("PLC", "118"): "CVL-OFI-2026/04965", ("PL", "2188"): "CVL-OFI-2026/05120",
    ("PL", "2199"): "CVL-OFI-2026/05115", ("PL", "2239"): "CVL-OFI-2026/05207",
    ("PL", "2242"): "CVL-OFI-2026/05372", ("PL", "2265"): "CVL-OFI-2026/05431",
    ("PL", "2276"): "CVL-OFI-2026/05446",
}

# Coautores = quem ALTEROU a peça (aprovação sem alteração não conta)
COAUTORES = {
    ("PL", "1883"): ["Marcel"], ("PL", "1884"): ["Marcel"], ("PL", "1900"): ["Marcel"],
    ("PL", "1934"): ["Marcel"], ("PLC", "98"): ["Marcel"], ("PL", "2040"): ["Luiza", "Marcel"],
    ("PL", "799"): ["Marcel"], ("PL", "2261"): ["Marcel"], ("PL", "2262"): ["Marcel"],
    ("PL", "2263"): ["Marcel"], ("PLC", "118"): ["Marcel"], ("PL", "2199"): ["Luiza", "Marcel"],
}


def peca_text(tipo, numero, ano):
    for suffix in ("-v2", ""):
        d = os.path.join(BASE, f"{tipo}-{numero}-{ano}{suffix}")
        if os.path.isdir(d):
            mds = glob.glob(os.path.join(d, "manifestacao_*short.md")) or glob.glob(
                os.path.join(d, "manifestacao_*.md"))
            if mds:
                with open(max(mds, key=os.path.getmtime), encoding="utf-8") as f:
                    return f.read()
    return ""


def ementa_oficial(tipo, numero, ano, fallback):
    """Extrai a cláusula 'que <ementa>' da abertura da peça; senão usa a do mapa."""
    text = peca_text(tipo, numero, ano).replace("*", "")
    m = re.search(r"Em atenção.{0,400}?que (.+?),\s*esta Subsecretaria", text, re.S)
    if m:
        ementa = re.sub(r"\s+", " ", m.group(1)).strip().upper()
        return f'{tipo} nº {numero}/{ano}, que "{ementa}"'
    return f'{tipo} nº {numero}/{ano} — "{re.sub(chr(10), " ", fallback).upper()}"'


def main():
    with open(MAP, encoding="utf-8") as f:
        rows = [r for r in csv.DictReader(f) if r["status"] == "Concluído"]
    print(f"Concluídos no mapa: {len(rows)}")

    new_rows = []
    for r in rows:
        tipo, num, ano = r["tipo"], r["numero"], r["ano"]
        resp = " + ".join(["Otávio"] + COAUTORES.get((tipo, num), []))
        obs = ementa_oficial(tipo, num, ano, r["ementa"])
        if (tipo, num) == ("PL", "163"):
            obs = 'Lei nº 9.326/2026 (PL nº 163/2025), que "INCLUI O DIA DO EMPREENDEDORISMO JOVEM NO CALENDÁRIO OFICIAL DA CIDADE" — Despacho de Ciência'
        if (tipo, num) == ("PL", "1320"):
            obs = 'Lei nº 9.337/2026 (PL nº 1320/2025), que "INCLUI O MÊS DO EMPREENDEDORISMO CARIOCA NO CALENDÁRIO OFICIAL DA CIDADE" — Despacho de Ciência'
        new_rows.append([OFICIOS[(tipo, num)], "-", resp, "CONCLUÍDO", obs])
        if (tipo, num) == ("PL", "1884"):  # 2ª rodada: autógrafo urgentíssimo
            new_rows.append([
                "CVL-OFI-2026/05305", "-", "Otávio + Luiza", "CONCLUÍDO",
                'Autógrafo do PL nº 1884/2026, que "INSTITUI E DISCIPLINA O CONCEITO DE '
                'EIXO ECONÔMICO MUNICIPAL" — ciência da sanção (texto idêntico ao já analisado)',
            ])
    print(f"Linhas a acrescentar (PL 1884 em 2 rodadas): {len(new_rows)}")

    if not os.path.exists(BACKUP):
        shutil.copy2(TARGET, BACKUP)
        print(f"Backup criado: {os.path.basename(BACKUP)}")

    wb = openpyxl.load_workbook(TARGET)
    ws = wb["SUBDEI"]

    # fim atual da tabela 19 (começa na linha 231; é a última tabela da aba)
    last = 231
    for row in range(232, ws.max_row + 1):
        if any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, 6)):
            last = row

    # dedup: nenhuma Ação nossa pode já existir na tabela
    existing = {str(ws.cell(row=row, column=1).value or "").strip() for row in range(232, last + 1)}
    dup = [nr[0] for nr in new_rows if nr[0] in existing]
    if dup:
        print("ABORTADO — ofícios já presentes na tabela:", dup, file=sys.stderr)
        sys.exit(1)

    template = [ws.cell(row=250, column=c) for c in range(1, 6)]
    for i, values in enumerate(new_rows):
        row = last + 1 + i
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            src = template[c - 1]
            cell.font = copy(src.font)
            cell.border = copy(src.border)
            cell.alignment = copy(src.alignment)
            cell.fill = copy(src.fill)
            cell.number_format = src.number_format

    wb.save(TARGET)
    print(f"Gravado: {TARGET} (linhas {last + 1}–{last + len(new_rows)})")

    wb2 = openpyxl.load_workbook(TARGET)
    ws2 = wb2["SUBDEI"]
    for i in range(len(new_rows)):
        row = last + 1 + i
        got = [ws2.cell(row=row, column=c).value for c in range(1, 6)]
        assert got == new_rows[i], f"linha {row} divergente"
    print(f"Verificação OK: {len(new_rows)} linhas relidas e conferidas.")


if __name__ == "__main__":
    main()
