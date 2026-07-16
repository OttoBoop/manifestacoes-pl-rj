#!/usr/bin/env python3
"""Mescla os despachos do time no PLs.xlsx (controle da gestão) com status de revisão humana.

Fluxo:
 1. Backup do PLs.xlsx original (uma vez).
 2. Valida a posição de cada peça do repo contra o mapa curado (controle-excel/revisao_map.csv).
 3. Acrescenta colunas Posição / Revisão humana / Revisor(es) / Lote/Entrega / Observações
    e as linhas do mapa na aba Página1, preservando as linhas da Adriana.
 4. Reabre o arquivo gravado e confere contagens.

A fonte da classificação de revisão é a transcrição completa do WhatsApp
("Conversa do WhatsApp com Despachos SUBDEI.txt") — evidência linha a linha no CSV.
"""

import csv
import glob
import os
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment, PatternFill

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CHAT_DIR = os.path.join(BASE, "conversa completa", "WhatsApp Chat with Despachos SUBDEI")
XLSX = os.path.join(CHAT_DIR, "PLs.xlsx")
BACKUP = os.path.join(CHAT_DIR, "PLs_backup_original_2026-07-16.xlsx")
MAP = os.path.join(BASE, "controle-excel", "revisao_map.csv")

NEW_HEADERS = ["Posição", "Revisão humana", "Revisor(es)", "Lote/Entrega", "Observações"]
HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFB4C6E7")
PENDING_FILL = PatternFill(fill_type="solid", fgColor="FFFFF2CC")


def find_manifest_md(tipo, numero, ano):
    """Localiza o .md final da peça no repo (preferindo -short.md e pastas -v2)."""
    for suffix in ("-v2", ""):
        d = os.path.join(BASE, f"{tipo}-{numero}-{ano}{suffix}")
        if not os.path.isdir(d):
            continue
        shorts = glob.glob(os.path.join(d, "manifestacao_*short.md"))
        if shorts:
            return max(shorts, key=os.path.getmtime)
        others = glob.glob(os.path.join(d, "manifestacao_*.md"))
        if others:
            return max(others, key=os.path.getmtime)
    return None


def classify_position(text):
    """Classifica a posição pela primeira marca no texto (padrão estável das peças)."""
    low = text.lower().replace("*", "")
    markers = {
        "contrario": ["sentido contrário", "posicionamento contrário"],
        "ciencia": ["toma ciência"],
        "nada a opor": ["nada a opor", "nada opor"],
    }
    first = (None, len(low) + 1)
    for label, terms in markers.items():
        for t in terms:
            i = low.find(t)
            if i != -1 and i < first[1]:
                first = (label, i)
    return first[0]


def validate(rows):
    problems = []
    for r in rows:
        expected = r["expected_repo"].strip()
        if not expected:
            continue
        md = find_manifest_md(r["tipo"], r["numero"], r["ano"])
        if md is None:
            problems.append(f"{r['tipo']} {r['numero']}/{r['ano']}: peça .md não encontrada no repo")
            continue
        with open(md, encoding="utf-8") as f:
            got = classify_position(f.read())
        if got != expected:
            problems.append(
                f"{r['tipo']} {r['numero']}/{r['ano']}: posição no repo = {got!r}, esperado {expected!r} ({md})"
            )
    if problems:
        print("VALIDAÇÃO FALHOU — nada foi gravado:", file=sys.stderr)
        for p in problems:
            print("  -", p, file=sys.stderr)
        sys.exit(1)
    print(f"Validação OK: {sum(1 for r in rows if r['expected_repo'].strip())} peças conferidas contra o repo.")


def last_data_row(ws):
    last = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=7):
        if any(c.value not in (None, "") for c in row):
            last = row[0].row
    return last


def main():
    with open(MAP, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    print(f"Mapa curado: {len(rows)} despachos.")

    validate(rows)

    if not os.path.exists(BACKUP):
        shutil.copy2(XLSX, BACKUP)
        print(f"Backup criado: {os.path.basename(BACKUP)}")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Página1"]
    last = last_data_row(ws)
    n_existing = last - 1
    print(f"Linhas existentes (Adriana): {n_existing} (até a linha {last}).")

    wrap = Alignment(wrap_text=True, vertical="top")

    # Cabeçalhos novos H1:L1, com o mesmo preenchimento do cabeçalho original
    for i, name in enumerate(NEW_HEADERS):
        c = ws.cell(row=1, column=8 + i, value=name)
        c.fill = HEADER_FILL

    # Linhas da Adriana: colunas novas ficam "—" (despachos anteriores à automação)
    for row in range(2, last + 1):
        for col in range(8, 13):
            ws.cell(row=row, column=col, value="—")

    # Linhas do time
    for j, r in enumerate(rows):
        row = last + 1 + j
        numero = int(r["numero"]) if r["numero"].isdigit() else r["numero"]
        values = [
            r["tipo"], numero, int(r["ano"]), r["responsavel"], r["status"], r["ementa"],
            r["posicao_final"], r["revisao_nivel"], r["revisores"], r["lote"],
            (r["obs"] + ("; " if r["obs"] and r["evidencia"] else "") + r["evidencia"]).strip("; "),
        ]
        for k, v in enumerate(values):
            c = ws.cell(row=row, column=2 + k, value=v)
            if k in (4, 5, 6, 7, 10):  # Status, Ementa, Posição, Revisão, Observações
                c.alignment = wrap
        if r["status"].startswith("Pendente"):
            ws.cell(row=row, column=6).fill = PENDING_FILL  # célula de Status

    for col, w in {"H": 34, "I": 24, "J": 22, "K": 13, "L": 70}.items():
        ws.column_dimensions[col].width = w

    wb.save(XLSX)
    print(f"Gravado: {XLSX}")

    # Verificação pós-gravação
    wb2 = openpyxl.load_workbook(XLSX)
    ws2 = wb2["Página1"]
    last2 = last_data_row(ws2)
    total = last2 - 1
    assert total == n_existing + len(rows), f"esperava {n_existing + len(rows)} linhas, achei {total}"
    keys = {(str(r["tipo"]), str(r["numero"]), str(r["ano"])) for r in rows}
    found = set()
    counts = {}
    for row in ws2.iter_rows(min_row=2, max_row=last2, min_col=2, max_col=12):
        key = (str(row[0].value), str(row[1].value), str(row[2].value))
        if key in keys:
            found.add(key)
        level = row[7].value
        counts[level] = counts.get(level, 0) + 1
    missing = keys - found
    assert not missing, f"linhas do mapa ausentes no xlsx: {missing}"
    print(f"Verificação OK: {total} linhas de dados ({n_existing} Adriana + {len(rows)} time).")
    print("Contagem por nível de revisão:")
    for level, n in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"  {level}: {n}")


if __name__ == "__main__":
    main()
